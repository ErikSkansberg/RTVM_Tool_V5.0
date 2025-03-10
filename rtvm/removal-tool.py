# rtvm/gui/tools/removal_tool.py - Tool for removing previously submitted change requests

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import logging
import os
from typing import Dict, List, Optional, Set, Tuple, Any
import queue
import threading

logger = logging.getLogger(__name__)

class RemovalTool:
    """
    Tool for removing previously submitted contractor proposed change requests.
    Compares the current RTVM file with an older submission and removes
    requests that have already been submitted.
    """
    
    def __init__(self, master: tk.Widget, app):
        """
        Initialize the removal tool.
        
        Args:
            master: The parent widget
            app: The main application instance
        """
        self.master = master
        self.app = app
        
        # Create and show the tool window
        self.create_window()
    
    def create_window(self):
        """Create the removal tool window."""
        self.window = tk.Toplevel(self.master)
        self.window.title("Remove Previously Submitted Contractor Proposed Change Request")
        self.window.geometry("800x400")
        self.window.transient(self.master)  # Set as transient to main window
        
        # Explanation Label at the top
        explanation = (
            "This tool compares the currently loaded file with an older submission.\n"
            "Any previously submitted contractor proposed change requests found in the old file\n"
            "will be removed from the new file upon clicking 'Remove Previously Submitted Requests'."
        )
        ttk.Label(self.window, text=explanation, justify="left").pack(pady=10, padx=10)
        
        # Frame for displaying current file and old file info
        files_frame = ttk.Frame(self.window, padding="10")
        files_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        
        # Label for the currently loaded new file
        ttk.Label(files_frame, text="Current (New) File:").grid(row=0, column=0, sticky="w")
        self.new_file_label_var = tk.StringVar(
            value=self.app.model.excel_file_path if self.app.model.excel_file_path else "No File Loaded"
        )
        ttk.Label(files_frame, textvariable=self.new_file_label_var).grid(row=0, column=1, sticky="w")
        
        # Label and button for uploading old file
        ttk.Label(files_frame, text="Old File:").grid(row=1, column=0, sticky="w")
        self.old_file_label_var = tk.StringVar(value="No Old File Selected")
        ttk.Label(files_frame, textvariable=self.old_file_label_var).grid(row=1, column=1, sticky="w")
        ttk.Button(
            files_frame, text="Browse Old File", command=self.browse_old_file
        ).grid(row=1, column=2, padx=5, sticky="w")
        
        # Button to remove previously submitted requests
        ttk.Button(
            self.window,
            text="Remove Previously Submitted Requests",
            command=self.remove_previously_submitted_requests
        ).pack(pady=20)
        
        # Create a frame for the progress bar (initially hidden)
        self.progress_frame = ttk.Frame(self.window, padding="10")
        self.progress_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        self.progress_frame.pack_forget()  # Hide initially
        
        ttk.Label(self.progress_frame, text="Processing...").pack(side=tk.LEFT, padx=5)
        self.removal_progress = ttk.Progressbar(
            self.progress_frame, orient='horizontal', length=300, mode='determinate'
        )
        self.removal_progress.pack(side=tk.LEFT, padx=5)
        
        # Results Text widget
        self.results_frame = ttk.LabelFrame(self.window, text="Results", padding="10")
        self.results_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.results_text = tk.Text(self.results_frame, height=10, width=80)
        self.results_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Add scrollbar for results text
        scrollbar = ttk.Scrollbar(self.results_frame, orient=tk.VERTICAL, command=self.results_text.yview)
        self.results_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(self.window, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
    
    def browse_old_file(self):
        """Browse for the old Excel file."""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xls;*.xlsx")]
        )
        if file_path:
            self.old_file_path = file_path
            self.old_file_label_var.set(os.path.basename(file_path))
    
    def remove_previously_submitted_requests(self):
        """Remove previously submitted change requests from the current file."""
        # Check if main new file and old file are uploaded
        if not hasattr(self.app.model, 'excel_file_path') or not self.app.model.excel_file_path:
            messagebox.showerror("Error", "No main (new) file is loaded. Please upload a main file first.")
            return
        
        if not hasattr(self, 'old_file_path') or not self.old_file_path:
            messagebox.showerror("Error", "No old file is selected.")
            return
        
        # Show progress bar
        self.progress_frame.pack()
        self.removal_progress['value'] = 0
        self.status_var.set("Loading old file...")
        self.window.update_idletasks()
        
        # Create a queue for thread communication
        progress_queue = queue.Queue()
        
        # Start the removal process in a separate thread
        threading.Thread(
            target=self._process_removal,
            args=(progress_queue,),
            daemon=True
        ).start()
        
        # Start monitoring the queue
        self._monitor_progress(progress_queue)
    
    def _process_removal(self, progress_queue):
        """
        Process the removal of previously submitted requests in a background thread.
        
        Args:
            progress_queue: Queue for thread communication
        """
        try:
            # Load the old Excel file
            old_df = pd.read_excel(self.old_file_path)
            progress_queue.put(('status', "Comparing files..."))
            
            # Identify columns in the current DF
            # Proposed Changes Column (G) is known to be at index 6
            proposed_changes_col = 6
            
            # Identify the "Contractor Proposed Change Comment Input" column by name
            comment_col_name = "Contractor Proposed Change Comment Input"
            if comment_col_name not in self.app.model.df.columns:
                progress_queue.put(('error', f"Column '{comment_col_name}' not found in new file."))
                return
            
            comment_col = self.app.model.df.columns.get_loc(comment_col_name)
            
            # Validate column indexes in the old file
            if proposed_changes_col >= len(self.app.model.df.columns):
                progress_queue.put(('error', "Proposed Changes column not found in new file."))
                return
            if proposed_changes_col >= len(old_df.columns):
                progress_queue.put(('error', "Proposed Changes column not found in old file."))
                return
            
            if comment_col >= len(self.app.model.df.columns):
                progress_queue.put(('error', f"'{comment_col_name}' column not found in new file."))
                return
            if comment_col >= len(old_df.columns):
                progress_queue.put(('error', f"'{comment_col_name}' column not found in old file."))
                return
            
            # Convert object ID columns to string (Object ID is assumed to be column 0)
            object_id_col = 0
            self.app.model.df.iloc[:, object_id_col] = self.app.model.df.iloc[:, object_id_col].astype(str)
            old_df.iloc[:, object_id_col] = old_df.iloc[:, object_id_col].astype(str)
            
            # Create a dictionary from old_df for old patterns
            progress_queue.put(('status', "Processing old patterns..."))
            old_patterns_map = {}
            
            for idx, row in old_df.iterrows():
                obj_id = str(row.iloc[object_id_col])
                old_patterns_str = row.iloc[proposed_changes_col]
                if pd.isna(old_patterns_str):
                    old_patterns_str = ""
                old_lines = [line.strip() for line in str(old_patterns_str).split('\n') if line.strip()]
                old_patterns_set = set(old_lines)
                
                old_patterns_map[obj_id] = {
                    'patterns': old_patterns_set
                }
            
            # Also create a dictionary for old comments (from the old file)
            old_comments_map = {}
            for idx, row in old_df.iterrows():
                obj_id = str(row.iloc[object_id_col])
                old_comments_str = row.iloc[comment_col]
                if pd.isna(old_comments_str):
                    old_comments_str = ""
                old_comment_lines = [line.strip() for line in str(old_comments_str).split('\n') if line.strip()]
                old_comments_set = set(old_comment_lines)
                
                if obj_id not in old_patterns_map:
                    old_patterns_map[obj_id] = {'patterns': set()}
                
                # Add comments to the same map for convenience
                old_patterns_map[obj_id]['comments'] = old_comments_set
            
            # Initialize progress bar
            total_rows = len(self.app.model.df)
            progress_queue.put(('max', total_rows))
            
            changes_made = False
            removed_patterns = []
            removed_comments = []
            
            # Remove previously submitted requests from Proposed Changes (column G)
            progress_queue.put(('status', "Removing patterns and comments..."))
            
            for idx, row in self.app.model.df.iterrows():
                progress_queue.put(('progress', idx + 1))
                
                obj_id = str(row.iloc[object_id_col])
                new_patterns_str = row.iloc[proposed_changes_col]
                if pd.isna(new_patterns_str):
                    new_patterns_str = ""
                new_lines = [line.strip() for line in str(new_patterns_str).split('\n') if line.strip()]
                
                if obj_id in old_patterns_map:
                    old_patterns_set = old_patterns_map[obj_id]['patterns']
                    # Filter out any patterns that exist in old file
                    filtered_lines = []
                    for line in new_lines:
                        if line in old_patterns_set:
                            removed_patterns.append(f"Row {idx+2}: {line}")
                        else:
                            filtered_lines.append(line)
                    
                    if len(filtered_lines) != len(new_lines):
                        changes_made = True
                        updated_str = '\n'.join(filtered_lines)
                        self.app.model.df.iat[idx, proposed_changes_col] = updated_str
                
                # Remove previously submitted comments
                new_comments_str = row.iloc[comment_col]
                if pd.isna(new_comments_str):
                    new_comments_str = ""
                new_comment_lines = [line.strip() for line in str(new_comments_str).split('\n') if line.strip()]
                
                if obj_id in old_patterns_map and 'comments' in old_patterns_map[obj_id]:
                    old_comments_set = old_patterns_map[obj_id]['comments']
                    filtered_comment_lines = []
                    for line in new_comment_lines:
                        if line in old_comments_set:
                            removed_comments.append(f"Row {idx+2}: {line}")
                        else:
                            filtered_comment_lines.append(line)
                    
                    if len(filtered_comment_lines) != len(new_comment_lines):
                        changes_made = True
                        updated_comments_str = '\n'.join(filtered_comment_lines)
                        self.app.model.df.iat[idx, comment_col] = updated_comments_str
            
            if not changes_made:
                progress_queue.put(('result', "No previously submitted requests or comments were found to remove."))
                return
            
            # If changes were made, update the Excel file
            progress_queue.put(('status', "Saving changes to Excel file..."))
            
            from openpyxl import load_workbook
            
            try:
                wb = load_workbook(self.app.model.excel_file_path)
                ws = wb.active  # Adjust if you need a specific sheet
                
                # Find the column indexes in the Excel sheet
                # Column G (Proposed Changes) = 7 in 1-based indexing
                # For 'Contractor Proposed Change Comment Input', find the correct column by header
                header_row = 1
                comment_col_letter = None
                for col in range(1, ws.max_column + 1):
                    header_val = ws.cell(row=header_row, column=col).value
                    if header_val == comment_col_name:
                        comment_col_letter = col
                        break
                
                # Update only the changed cells
                for idx, row in enumerate(self.app.model.df.iterrows()):
                    row_index, row_data = row
                    new_content = row_data.iloc[proposed_changes_col]
                    if pd.isna(new_content):
                        new_content = ""
                    ws.cell(row=idx+2, column=7, value=new_content)  # Proposed Changes column (G)
                    
                    new_comment_content = row_data.iloc[comment_col]
                    if pd.isna(new_comment_content):
                        new_comment_content = ""
                    if comment_col_letter is not None:
                        ws.cell(row=idx+2, column=comment_col_letter, value=new_comment_content)
                
                wb.save(self.app.model.excel_file_path)
                
                # Prepare result message
                result_message = (
                    f"Successfully removed {len(removed_patterns)} previously submitted patterns and "
                    f"{len(removed_comments)} previously submitted comments.\n\n"
                    "Removed patterns:\n" + "\n".join(removed_patterns[:20])
                )
                
                if len(removed_patterns) > 20:
                    result_message += f"\n... and {len(removed_patterns) - 20} more"
                
                result_message += "\n\nRemoved comments:\n" + "\n".join(removed_comments[:20])
                
                if len(removed_comments) > 20:
                    result_message += f"\n... and {len(removed_comments) - 20} more"
                
                progress_queue.put(('result', result_message))
                
            except Exception as e:
                logger.error(f"Failed to save updated file: {e}", exc_info=True)
                progress_queue.put(('error', f"Failed to save updated file: {str(e)}"))
        
        except Exception as e:
            logger.error(f"Error in removal process: {e}", exc_info=True)
            progress_queue.put(('error', f"An unexpected error occurred: {str(e)}"))
    
    def _monitor_progress(self, queue):
        """
        Monitor progress updates from the background thread.
        
        Args:
            queue: Queue containing progress updates
        """
        try:
            # Check for messages in the queue
            try:
                message_type, data = queue.get_nowait()
                
                if message_type == 'status':
                    self.status_var.set(data)
                
                elif message_type == 'max':
                    self.removal_progress['maximum'] = data
                
                elif message_type == 'progress':
                    self.removal_progress['value'] = data
                    # Update progress percentage
                    if self.removal_progress['maximum'] > 0:
                        percent = int((data / self.removal_progress['maximum']) * 100)
                        self.status_var.set(f"Processing... {percent}%")
                
                elif message_type == 'result':
                    # Hide progress bar
                    self.progress_frame.pack_forget()
                    
                    # Show result in the text widget
                    self.results_text.delete(1.0, tk.END)
                    self.results_text.insert(tk.END, data)
                    
                    # Update status
                    self.status_var.set("Completed")
                    
                    # Update the app's UI to reflect changes
                    self.app.update_ui_after_navigation()
                    
                    # Show success message
                    messagebox.showinfo("Success", "Previously submitted requests and comments have been removed.")
                    
                    return  # Done monitoring
                
                elif message_type == 'error':
                    # Hide progress bar
                    self.progress_frame.pack_forget()
                    
                    # Show error in the text widget
                    self.results_text.delete(1.0, tk.END)
                    self.results_text.insert(tk.END, f"ERROR: {data}")
                    
                    # Update status
                    self.status_var.set("Error")
                    
                    # Show error message
                    messagebox.showerror("Error", data)
                    
                    return  # Done monitoring
            
            except queue.Empty:
                pass  # No messages in the queue
            
            # Schedule the next check
            self.window.after(100, lambda: self._monitor_progress(queue))
            
        except Exception as e:
            logger.error(f"Error monitoring progress: {e}", exc_info=True)
            messagebox.showerror("Error", f"An unexpected error occurred while monitoring progress: {str(e)}")
